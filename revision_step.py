import pandas as pd
from datetime import timedelta
from openpyxl import load_workbook

"""
Funciones para realizar la revisión de archivos Excel relacionados con la publicidad y los spots.
Estas funciones incluyen la eliminación de filas obsoletas, la revisión de "Back to back", la comparación de spots con la pauta y la revisión de creativos.
además, se calcula el resultado final de la revisión.
"""

def delete_outdated_rows(final_path, start_date, end_date, sheet_name):
    """
    Elimina las filas de un archivo Excel que están fuera de un rango de fechas específico.
    Se utiliza la columna 'Date Time Zone' para determinar si una fila está dentro del rango.
    """
    # Leer la hoja específica del archivo Excel
    df = pd.read_excel(final_path, sheet_name=sheet_name)

    # Convertir 'Date Time Zone' a formato datetime (incluyendo hora)
    df['Date Time Zone'] = pd.to_datetime(df['Date Time Zone'])

    # Crear una columna auxiliar con solo la fecha (sin hora)
    df['aux_date'] = df['Date Time Zone'].dt.date

    # Convertir start_date y end_date a formato datetime.date
    start_date = pd.to_datetime(start_date).date()
    end_date = pd.to_datetime(end_date).date()

    # Filtrar filas dentro del rango de fechas
    #df = df[(df['aux_date'] >= start_date) & (df['aux_date'] <= end_date)]
    df = df[(df['aux_date'] >= start_date)]

    # Eliminar la columna auxiliar
    df.drop(columns=['aux_date'], inplace=True)
    
    df['Date Time Zone'] = pd.to_datetime(df['Date Time Zone']).dt.strftime('%m/%d/%Y %H:%M:%S')

    # Guardar los datos actualizados en la hoja específica
    with pd.ExcelWriter(final_path, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)

def remove_not_found_rows(final_path, sheet_name):
    """
    Elimina las filas que tienen el estado 'Not Found' en la columna 'Estado' y las mueve a una nueva hoja llamada 'No encontrados'.
    Si la hoja ya existe, se elimina y se crea una nueva.
    """
    nf_sheet_name='No encontrados'
            
    # Leer la hoja específica del archivo Excel
    df = pd.read_excel(final_path, sheet_name=sheet_name)
    #Solo si la columna Estado contiene 'Not Found' se elimina la fila, en caso de que no tenga ningun valor se pasa
    if 'Estado' in df.columns:
        if 'Not Found' in df['Estado'].values:
            
            #Abrir con openpyxl para poder agregar hojas
            wb = load_workbook(final_path)
            #Crear una hoja nueva con el  nombre 'No encontrados'
            if nf_sheet_name in wb.sheetnames:
                wb.remove(wb[nf_sheet_name])
            ws = wb.create_sheet(title=nf_sheet_name)
            
            # mover las filas que contienen en la columna Estado = "Not Found" a una nueva hoja llamada 'No encontrados' y dejar el final path en la hoja sheet_name sin esas filas
            not_found_df = df[df['Estado'] == 'Not Found']
            df = df[df['Estado'] != 'Not Found']
    
            with pd.ExcelWriter(final_path, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)
            with pd.ExcelWriter(final_path, mode='a', engine='openpyxl', if_sheet_exists='overlay') as writer:
                not_found_df.to_excel(writer, sheet_name=nf_sheet_name, index=False)
        else:
            print('No rows with "Not Found" status found')
    else:
        print('No "Estado" column found')

def b2bV2(final_path, sheet_name):
    
    """
    Realiza la revisión de "Back to back" en un archivo Excel específico.
    Se basa en la columna 'Feed Index' y 'Date Time Zone' para determinar si los registros cumplen con la condición de "Back to back".
    """
    # Leer archivo Excel y hoja
    df = pd.read_excel(final_path, sheet_name=sheet_name)

    # Convertir 'Date Time Zone' a tipo datetime
    df['Date Time Zone'] = pd.to_datetime(df['Date Time Zone'], errors='coerce')

    # Ordenar datos por 'Feed Index' y 'Date Time Zone'
    df.sort_values(by=['Feed Index', 'Date Time Zone'], inplace=True)
    df.reset_index(drop=True, inplace=True)

    # Inicializar columna 'Back to back'
    df['Back to back'] = ''
    
    # Variables para almacenar los valores previos
    previous_feed_index = None
    previous_date = None
    previous_duration = None

    # Iterar sobre las filas del DataFrame
    for idx, row in df.iterrows():
        current_feed_index = row['Feed Index']
        current_date = row['Date Time Zone']
        current_duration = row['Duracion']

        # Validar datos nulos o inválidos
        if pd.isna(current_date) or pd.isna(current_duration):
            df.at[idx, 'Back to back'] = 'Error: Datos incompletos'
            continue
        if not isinstance(current_duration, (int, float)):
            df.at[idx, 'Back to back'] = 'Error: Duración inválida'
            continue

        if idx == 0:  # Primera fila
            df.at[idx, 'Back to back'] = 'Ok'
        else:
            # Calcular la fecha límite para "Back to back"
            date_with_seconds = previous_date + timedelta(seconds=previous_duration + 2)

            # Verificar si cumple la condición de "Back to back"
            if current_feed_index == previous_feed_index and current_date <= date_with_seconds:
                df.at[idx, 'Back to back'] = 'Back to back'
            else:
                df.at[idx, 'Back to back'] = 'Ok'

        # Actualizar las variables previas
        previous_feed_index = current_feed_index
        previous_date = current_date
        previous_duration = current_duration

    # Guardar el DataFrame actualizado en el archivo Excel
    with pd.ExcelWriter(final_path, mode='a', engine='openpyxl', if_sheet_exists='overlay') as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        
def rev_spots_vs_pauta(final_path, filtered_bdd_path, sheet_name):
    """
    Realiza la revisión de los spots en comparación con la pauta.
    Se basa en la columna 'Feed Index', 'Brand' y 'Date Time Zone' para determinar si los registros cumplen con la pauta.
    """          
    # Leer los datos desde los archivos Excel
    df = pd.read_excel(final_path, sheet_name=sheet_name)
    df_bdd = pd.read_excel(filtered_bdd_path)
    
    # Crear y preparar columnas necesarias
    df_bdd['Spot status'] = ''  # Inicializar el estado de los spots
    df['Revision type'] = pd.to_numeric(df['Revision type'], errors='coerce').fillna(0).astype(int)
    df['Rev vs pauta'] = ''
    df['Spot Observation'] = ''
    df['Fecha Final Revision'] = ''
    df['Rate'] = ''
    
    # Asegurarse de que las fechas están en formato datetime
    df_bdd['Date Time Zone'] = pd.to_datetime(df_bdd['Date Time Zone'], format='%m/%d/%Y %H:%M:%S')
    
    for idx, row in df.iterrows():
        current_feed_index = row['Feed Index']
        current_brand = row['Brand']
        
        # Filtrar registros correspondientes en df_bdd
        matches = df_bdd[(df_bdd['Feed Index'] == current_feed_index) & (df_bdd['Brand'] == current_brand)]
        
        if not matches.empty:
            for idx2, row2 in matches.iterrows():
                # Verificar si el registro ya fue procesado
                if row2['Spot status'] == 'Ok':
                    continue
                
                if row['Revision type'] == 1:
                    # Procesamiento para "Revision type" 1
                    current_minus_date = pd.to_datetime(row['Date Time Zone - Minutes'], errors='coerce')
                    current_equal_date = pd.to_datetime(row['Date Time Zone = Minutes'], errors='coerce')
                    current_plus_date = pd.to_datetime(row['Date Time Zone + Minutes'], errors='coerce')
                    comparer_date = pd.to_datetime(row2['Date Time Zone'], errors='coerce')
                    current_rate = row2['Rate']
                    
                    if current_minus_date == comparer_date:
                        df.at[idx, 'Rev vs pauta'] = 'Ok'
                        df.at[idx, 'Fecha Final Revision'] = f'{current_minus_date.strftime("%m/%d/%Y %H:%M:%S")}'
                        df.at[idx, 'Spot Observation'] = 'Spot Correcto'
                        df_bdd.at[idx2, 'Spot status'] = 'Ok'
                        df.at[idx, 'Rate'] = current_rate
                        break
                    if current_equal_date == comparer_date:
                        df.at[idx, 'Rev vs pauta'] = 'Ok'
                        df.at[idx, 'Fecha Final Revision'] = f'{current_equal_date.strftime("%m/%d/%Y %H:%M:%S")}'
                        df.at[idx, 'Spot Observation'] = 'Spot Correcto'
                        df_bdd.at[idx2, 'Spot status'] = 'Ok'
                        df.at[idx, 'Rate'] = row2['Rate']
                        break
                    if current_plus_date == comparer_date:
                        df.at[idx, 'Rev vs pauta'] = 'Ok'
                        df.at[idx, 'Fecha Final Revision'] = f'{current_plus_date.strftime("%m/%d/%Y %H:%M:%S")}'
                        df.at[idx, 'Spot Observation'] = 'Spot Correcto'
                        df_bdd.at[idx2, 'Spot status'] = 'Ok'
                        df.at[idx, 'Rate'] = row2['Rate']
                        break
                elif row['Revision type'] == 2:
                    # Procesamiento para "Revision type" 2
                    current_minus_dp = row['Day Part - Minutes']
                    current_equal_dp = row['Day Part = Minutes']
                    current_plus_dp = row['Day Part + Minutes']
                    
                    current_minus_full_day = pd.to_datetime(row['Full Day - Minutes'], errors='coerce')
                    current_equal_full_day = pd.to_datetime(row['Full Day = Minutes'], errors='coerce')
                    current_plus_full_day = pd.to_datetime(row['Full Day + Minutes'], errors= 'coerce')
                    
                    comparer_date_full_day = pd.to_datetime(row2['Date Full Day'], errors='coerce')
                    comparer_day_part = row2['Franja']
                    current_rate = row2['Rate']
                    
                    if  current_minus_dp.strip().lower() == row2['Franja'].strip().lower() and current_minus_full_day == comparer_date_full_day:
                        df.at[idx, 'Rev vs pauta'] = 'Ok'
                        df.at[idx, 'Fecha Final Revision'] = f'{current_minus_dp}'
                        df.at[idx, 'Spot Observation'] = 'Spot Correcto'
                        df_bdd.at[idx2, 'Spot status'] = 'Ok'
                        df.at[idx, 'Rate'] = current_rate
                        break
                    
                    if current_equal_dp.strip().lower() == row2['Franja'].strip().lower() and current_equal_full_day == comparer_date_full_day:
                        df.at[idx, 'Rev vs pauta'] = 'Ok'
                        df.at[idx, 'Fecha Final Revision'] = f'{current_equal_dp}'
                        df.at[idx, 'Spot Observation'] = 'Spot Correcto'
                        df_bdd.at[idx2, 'Spot status'] = 'Ok'
                        df.at[idx, 'Rate'] = current_rate
                        break
                    if current_plus_dp.strip().lower() == row2['Franja'].strip().lower() and current_plus_full_day == comparer_date_full_day:
                        df.at[idx, 'Rev vs pauta'] = 'Ok'
                        df.at[idx, 'Fecha Final Revision'] = f'{current_plus_dp}'
                        df.at[idx, 'Spot Observation'] = 'Spot Correcto'
                        df_bdd.at[idx2, 'Spot status'] = 'Ok'
                        df.at[idx, 'Rate'] = current_rate
                        break

                elif row['Revision type'] == 3:
                    # Procesamiento para "Revision type" 3
                    current_minus_full_day = pd.to_datetime(row['Full Day - Minutes'], format='%m/%d/%Y %H:%M')
                    current_equal_full_day = pd.to_datetime(row['Full Day = Minutes'], format='%m/%d/%Y %H:%M')
                    current_plus_full_day = pd.to_datetime(row['Full Day + Minutes'], format='%m/%d/%Y %H:%M')
                    comparer_date_full_day = pd.to_datetime(row2['Date Full Day'], errors='coerce')
                    current_rate = row2['Rate']
                    
                    if current_minus_full_day == comparer_date_full_day:
                        df.at[idx, 'Rev vs pauta'] = 'Ok'
                        df.at[idx, 'Fecha Final Revision'] = f'{current_minus_full_day.strftime("%m/%d/%Y %H:%M")}'
                        df.at[idx, 'Spot Observation'] = 'Spot Correcto'
                        df_bdd.at[idx2, 'Spot status'] = 'Ok'
                        df.at[idx, 'Rate'] = current_rate
                        break
                    elif current_equal_full_day == comparer_date_full_day:
                        df.at[idx, 'Rev vs pauta'] = 'Ok'
                        df.at[idx, 'Fecha Final Revision'] = f'{current_equal_full_day.strftime("%m/%d/%Y %H:%M")}'
                        df.at[idx, 'Spot Observation'] = 'Spot Correcto'
                        df_bdd.at[idx2, 'Spot status'] = 'Ok'
                        df.at[idx, 'Rate'] = current_rate
                        break
                    elif current_plus_full_day == comparer_date_full_day:
                        df.at[idx, 'Rev vs pauta'] = 'Ok'
                        df.at[idx, 'Fecha Final Revision'] = f'{current_plus_full_day.strftime("%m/%d/%Y %H:%M")}'
                        df.at[idx, 'Spot Observation'] = 'Spot Correcto'
                        df_bdd.at[idx2, 'Spot status'] = 'Ok'
                        df.at[idx, 'Rate'] = current_rate
                        break
    
    for idx, row in df.iterrows():
        
        #filtrar df por aquellos que aun no han sido revisados 
        if row['Rev vs pauta'] == '':
            current_feed_index = row['Feed Index']
            current_brand = row['Brand']
            
            matches = df_bdd[(df_bdd['Feed Index'] == current_feed_index) & (df_bdd['Brand'] == current_brand)]
            
            if not matches.empty:
                for idx2, row2 in matches.iterrows():
                    if not row2['Spot status'] == 'Ok':
                        continue
                    
                    if row['Revision type'] == 1:
                        current_minus_date = pd.to_datetime(row['Date Time Zone - Minutes'], errors='coerce')
                        current_equal_date = pd.to_datetime(row['Date Time Zone = Minutes'], errors='coerce')
                        current_plus_date = pd.to_datetime(row['Date Time Zone + Minutes'], errors='coerce')
                        comparer_date = pd.to_datetime(row2['Date Time Zone'], errors='coerce')
                        
                        if comparer_date in [current_minus_date, current_equal_date, current_plus_date]:
                            df.at[idx, 'Rev vs pauta'] = 'No'
                            df.at[idx, 'Spot Observation'] = 'Spot Duplicado'
                            df.at[idx, 'Rate'] = 0
                            
                            if comparer_date == current_minus_date:
                                df.at[idx, 'Fecha Final Revision'] = f'{current_minus_date.strftime("%m/%d/%Y %H:%M:%S")}'
                            elif comparer_date == current_equal_date:
                                df.at[idx, 'Fecha Final Revision'] = f'{current_equal_date.strftime("%m/%d/%Y %H:%M:%S")}'
                            elif comparer_date == current_plus_date:
                                df.at[idx, 'Fecha Final Revision'] = f'{current_plus_date.strftime("%m/%d/%Y %H:%M:%S")}'
                            break
                    if row['Revision type'] == 2:
                        current_minus_dp = row['Day Part - Minutes'].strip().lower()
                        current_equal_dp = row['Day Part = Minutes'].strip().lower()
                        current_plus_dp = row['Day Part + Minutes'].strip().lower()
                        
                        current_minus_full_day = pd.to_datetime(row['Full Day - Minutes'], errors='coerce')
                        current_equal_full_day = pd.to_datetime(row['Full Day = Minutes'], errors='coerce')
                        current_plus_full_day = pd.to_datetime(row['Full Day + Minutes'], errors='coerce')
                        
                        comparer_date_full_day = pd.to_datetime(row2['Date Full Day'], errors='coerce')
                        comparer_day_part = row2['Franja'].strip().lower()
                        
                        if comparer_date_full_day in [current_minus_full_day, current_equal_full_day, current_plus_full_day]:
                            if comparer_day_part in [current_minus_dp, current_equal_dp, current_plus_dp]:
                                df.at[idx, 'Rev vs pauta'] = 'No'
                                df.at[idx, 'Spot Observation'] = 'Spot Duplicado'
                                df.at[idx, 'Rate'] = 0
                                
                                if comparer_day_part == current_minus_dp:
                                    df.at[idx, 'Fecha Final Revision'] = f'{current_minus_dp}'
                                elif comparer_day_part == current_equal_dp:
                                    df.at[idx, 'Fecha Final Revision'] = f'{current_equal_dp}'
                                elif comparer_day_part == current_plus_dp:
                                    df.at[idx, 'Fecha Final Revision'] = f'{current_plus_dp}'
                                break
                    
                    if row['Revision type'] == 3:
                        current_minus_full_day = pd.to_datetime(row['Full Day - Minutes'], format='%m/%d/%Y %H:%M')
                        current_equal_full_day = pd.to_datetime(row['Full Day = Minutes'], format='%m/%d/%Y %H:%M')
                        current_plus_full_day = pd.to_datetime(row['Full Day + Minutes'], format='%m/%d/%Y %H:%M')
                        
                        comparer_date_full_day = pd.to_datetime(row2['Date Full Day'], errors='coerce')
                        
                        if comparer_date_full_day in [current_minus_full_day, current_equal_full_day, current_plus_full_day]:
                            df.at[idx, 'Rev vs pauta'] = 'No'
                            df.at[idx, 'Spot Observation'] = 'Spot Duplicado'
                            df.at[idx, 'Rate'] = 0
                            
                            if comparer_date_full_day == current_minus_full_day:
                                df.at[idx, 'Fecha Final Revision'] = f'{current_minus_full_day.strftime("%m/%d/%Y %H:%M")}'
                            elif comparer_date_full_day == current_equal_full_day:
                                df.at[idx, 'Fecha Final Revision'] = f'{current_equal_full_day.strftime("%m/%d/%Y %H:%M")}'
                            elif comparer_date_full_day == current_plus_full_day:
                                df.at[idx, 'Fecha Final Revision'] = f'{current_plus_full_day.strftime("%m/%d/%Y %H:%M")}'
                            break                        
       
        # Si no se encontró un registro válido
        if df.at[idx, 'Rev vs pauta'] == '':
            df.at[idx, 'Rev vs pauta'] = 'No'
            #asignar el valor de revision final como el valor = minutes perteneciente a la fila
            df.at[idx, 'Fecha Final Revision'] = pd.to_datetime(df.at[idx, 'Date Time Zone = Minutes']).strftime('%m/%d/%Y %H:%M:%S')
            df.at[idx, 'Spot Observation'] = 'Spot No solicitado'
            df.at[idx, 'Rate'] = 0
            
            
    
    df_bdd['Spot status'] = df_bdd['Spot status'].replace('', 'No')
    
    wb = load_workbook(final_path)
    sn = 'BDD Final Revisada'
    
    if sn in wb.sheetnames:
        wb.remove(wb[sn])
    
    ws = wb.create_sheet(title=sn)
    
    df_bdd['Date Time Zone']= df_bdd['Date Time Zone'].dt.strftime('%m/%d/%Y %H:%M:%S')
    df['Date Time Zone'] = pd.to_datetime(df['Date Time Zone'], errors='coerce')
    df['Date Time Zone'] = df['Date Time Zone'].dt.strftime('%m/%d/%Y %H:%M:%S')
    
    with pd.ExcelWriter(final_path, mode='a', engine='openpyxl', if_sheet_exists='overlay') as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        df_bdd.to_excel(writer, sheet_name=sn, index=False)

def rev_creatives(aux_path, final_file):
    """
    Realiza la revisión de los creativos en comparación con un archivo auxiliar.
    Se basa en la columna 'Rotation Id', 'Creativo' y 'Brand' para determinar si los registros cumplen con la pauta.
    """
    sheet_name = 'Archivo Final Play Logger'
    aux_sheet_name = 'Month_Rotation'
    
    aux_file = pd.read_excel(aux_path, sheet_name=aux_sheet_name)
    main_file = pd.read_excel(final_file, sheet_name=sheet_name)
    
    aux_file['Key'] = ''
    main_file['Key'] = ''
    main_file['Rev Creativos'] = ''
    main_file['Creative observation'] = ''
    main_file['Id Rev % Ads'] = ''
    main_file['Id Fecha Ads'] = ''
    
    #Convert columns to datetime with HH:MM:SS format
    aux_file['Start date'] = pd.to_datetime(aux_file['Start date'])
    aux_file['End date'] = pd.to_datetime(aux_file['End date'])
    main_file['Date Time Zone'] = pd.to_datetime(main_file['Date Time Zone'])
        
    #create a key column in the aux file
    # Crear la columna de clave utilizando operaciones vectorizadas
    aux_file['Key'] = aux_file['Rotation Id'].astype(str) + '_' + aux_file['Creativo'] + '_' + aux_file['Brand']

    #create a key column in the main file
    main_file['Key'] = main_file['Rotation Id'].astype(str) + '_' + main_file['Creativo'] + '_' + main_file['Brand']
    
    for idx, row in main_file.iterrows():
        if row['Estado'] == 'Found':
            # Obtener la clave y la fecha del archivo principal
            key = row['Key']
            date_rev = row['Date Time Zone']
            
            # Filtrar el archivo auxiliar por la clave
            aux_rows = aux_file[aux_file['Key'] == key]
            
            if not aux_rows.empty:
                for _, aux_row in aux_rows.iterrows():
                    # Comprobar si la fecha está en el rango
                    if pd.notnull(aux_row['Start date']) and pd.notnull(aux_row['End date']) and aux_row['Start date'] <= date_rev <= aux_row['End date']:
                        main_file.loc[idx, 'Rev Creativos'] = 'OK'
                        main_file.loc[idx, 'Creative observation'] = "Creativo Correcto"
                        main_file.loc[idx, 'Id Rev % Ads'] = aux_row['Id Rev % Ads']
                        main_file.loc[idx, 'Id Fecha Ads'] = aux_row['Id Fecha Ads']
                        break
                    else:
                        main_file.loc[idx, 'Rev Creativos'] = 'NO'
                        main_file.loc[idx, 'Creative observation'] = 'Creativo transmitido incorrectamente'
                        main_file.loc[idx, 'Id Rev % Ads'] = ''
                        main_file.loc[idx, 'Id Fecha Ads'] = ''
            else:
                main_file.loc[idx, 'Rev Creativos'] = 'NO'
                main_file.loc[idx, 'Creative observation'] = f"Creativo incorrecto"
                main_file.loc[idx, 'Id Rev % Ads'] = ''
                main_file.loc[idx, 'Id Fecha Ads'] = ''
        else:
            main_file.loc[idx, 'Rev Creativos'] = 'NO'
            main_file.loc[idx, 'Creative observation'] = 'Creativo incorrecto'
            main_file.loc[idx, 'Id Rev % Ads'] = ''
            main_file.loc[idx, 'Id Fecha Ads'] = ''

    
    #Drop key columns in both files
    aux_file.drop(columns=['Key'], inplace=True)
    main_file.drop(columns=['Key'], inplace=True)
    
    main_file['Date Time Zone'] = main_file['Date Time Zone'].dt.strftime('%m/%d/%Y %H:%M:%S')
    
    #Save the main file with the new columns
    with pd.ExcelWriter(final_file, mode='a', engine='openpyxl', if_sheet_exists='overlay') as writer:
        main_file.to_excel(writer, sheet_name=sheet_name, index=False)
    
def final_result(final_path):
    """
    Calcula el resultado final de la revisión comparando los resultados de las revisiones de pauta y creativos.
    Si ambos son 'Ok', el resultado es 'Ok', de lo contrario es 'No'.
    """
    #takes revision fields and compares, if every revision is Ok the result is ok, if one of this or both are No then the value is No
    sheet_name = 'Archivo Final Play Logger'
    main_file = pd.read_excel(final_path, sheet_name=sheet_name)
    
    main_file['Final Result'] = ''
    
    for idx, row in main_file.iterrows():
        if row['Rev vs pauta'] == 'Ok' and row['Rev Creativos'] == 'OK':
            main_file.loc[idx, 'Final Result'] = 'Ok'
        else:
            main_file.loc[idx, 'Final Result'] = 'No'
    
    with pd.ExcelWriter(final_path, mode='a', engine='openpyxl', if_sheet_exists='overlay') as writer:
        main_file.to_excel(writer, sheet_name=sheet_name, index=False)
                      
def full_revision(final_path, filtered_bdd_path, aux_path, start_date, end_date, sheet_name, log_func=None):
    """
    Realiza una revisión completa de los archivos de Play Logger y BDD.
    Elimina filas obsoletas, realiza revisiones de "Back to back", "Spot" y "Creativos", y calcula el resultado final.
    """
    if log_func: log_func("🧹 Eliminando filas obsoletas...")
    delete_outdated_rows(final_path, start_date, end_date, sheet_name)

    if log_func: log_func("🗂️ Eliminando filas no encontradas...")
    remove_not_found_rows(final_path, sheet_name)

    if log_func: log_func("🔄 Iniciando revisión Back to Back...")
    b2bV2(final_path, sheet_name)
    if log_func: log_func("✅ Revisión Back to Back completada")

    if log_func: log_func("🔎 Iniciando revisión de Spots...")
    rev_spots_vs_pauta(final_path, filtered_bdd_path, sheet_name)
    if log_func: log_func("✅ Revisión de Spots completada")

    if log_func: log_func("🎨 Iniciando revisión de Creativos...")
    rev_creatives(aux_path, final_path)
    if log_func: log_func("✅ Revisión de Creativos completada")

    if log_func: log_func("📊 Calculando resultado final...")
    final_result(final_path)
    if log_func: log_func("🏁 Revisión completa.")