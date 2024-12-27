import os
import openpyxl
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
import dotenv

dotenv.load_dotenv()

def apply_transformations_to_excel_file(excel_path):
    """
    Formatea un archivo de Excel específico ('Ads_Played_Log_per_Day.xlsx') aplicando varias transformaciones a sus celdas:
    
    1. Reemplaza caracteres Unicode U+00A0 por celdas vacías.
    2. Rellena celdas vacías con el valor de la celda superior.
    3. Reemplaza fórmulas en celdas por el último valor conocido.
    4. Elimina filas que contienen la palabra 'Conteo' en la columna 'Días laborados' (columna C).
    5. Convierte el formato de las fechas en una columna nueva 'F' usando el formato MM/DD/YYYY.
    
    Finalmente, guarda el archivo Excel con todos los cambios aplicados.
    """
    # Cargar el archivo Excel y seleccionar la primera hoja
    wb = load_workbook(excel_path)
    ws = wb.active

    # Aplicar las funciones de transformación
    print("Borrando unicode chars")
    replace_unicode_character(ws)

    print("Llenando celdas")
    fill_empty_cells(ws)

    print("Reemplazar formulas con valores")
    replace_formulas_with_values(ws)

    print("Limpiando tabla")
    delete_count_rows(ws)

    print("Formateando fechas")
    format_date_column(excel_path)
    
    # Guardar el archivo con los cambios
    wb.save(excel_path)
    wb.close()


#===============LOGIC BETWEEN TRANFORMATION===============#

def replace_unicode_character(ws):
    """Reemplaza el carácter Unicode U+00A0 por celdas vacías en toda la hoja de trabajo."""
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == '\u00A0':
                cell.value = None

def fill_empty_cells(ws):
    """
    Rellena celdas vacías en las primeras cinco columnas de la hoja
    con el valor de la celda inmediata superior, utilizando una fórmula.
    """
    for col in ws.iter_cols(min_col=1, max_col=5):
        for row in range(2, ws.max_row + 1):
            if col[row-1].value is None:
                cell_above = col[row-2]
                if cell_above.value is not None:
                    col[row-1].value = cell_above.value

def replace_formulas_with_values(ws):
    """
    Reemplaza las fórmulas en las primeras cinco columnas de la hoja con el último valor conocido,
    dejando el valor fijo en la celda sin fórmula.
    """
    for col in ws.iter_cols(min_col=1, max_col=5):
        last_known_value = None
        for row in range(1, ws.max_row + 1):
            cell = col[row-1]
            if cell.value is not None and cell.data_type != 'f':  # Check if it's not a formula
                last_known_value = cell.value
            elif cell.data_type == 'f':  # Replace formula with last known value
                cell.value = last_known_value

def delete_count_rows(ws):
    """
    Elimina las filas que contienen la palabra 'Conteo' en la tercera columna ('Días laborados').
    Recorre desde la última fila hasta la primera para evitar problemas de desplazamiento.
    """
    column_index = 3  # Column 'Días laborados'
    rows_to_delete = []
    for row in range(1, ws.max_row + 1):
        cell_value = ws.cell(row=row, column=column_index).value
        if cell_value and "Conteo" in str(cell_value):
            rows_to_delete.append(row)
    
    for row in reversed(rows_to_delete):
        ws.delete_rows(row)

def format_date_column(excel_path):

    sheet_name = 'Archivo Final Play Logger'

    try:
        df = pd.read_excel(excel_path, usecols='B', names=['Fecha'])
    except Exception as e:
        print(f"Error al leer el archivo Excel: {e}")
        return
    
    # Convertir fechas al formato MM/DD/YYYY
    df['Fecha'] = pd.to_datetime(df['Fecha'], format='%d/%m/%Y', errors='coerce').dt.strftime('%m/%d/%Y')

    # Cargar el archivo Excel y verificar la existencia de la hoja
    wb = load_workbook(excel_path)
    if sheet_name not in wb.sheetnames:
        wb.create_sheet(title=sheet_name)
        print(f"Hoja '{sheet_name}' creada.")
    ws = wb[sheet_name]

    # Escribir las fechas formateadas en la columna 'F'
    for idx, date_value in enumerate(df['Fecha'].fillna('Fecha Inválida'), start=1):
        ws.cell(row=idx, column=6, value=date_value)  # Columna 6 corresponde a 'F'

    # Guardar el archivo
    try:
        wb.save(excel_path)
        print(f"Archivo guardado correctamente en {excel_path}")
    except Exception as e:
        print(f"Error al guardar el archivo Excel: {e}")
    finally:
        wb.close()
