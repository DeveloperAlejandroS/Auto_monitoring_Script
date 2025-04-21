import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment, NamedStyle, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
import os

"""
Este script se encarga de generar un reporte final a partir de datos de un archivo auxiliar y un archivo de datos finales.
El reporte incluye un resumen por proveedor, detalles de cada proveedor y tablas de rotación.
Adema se generan hojas de programación para cada proveedor y se aplican estilos a las tablas generadas.
"""

def generate_report(aux_path):
    """
    Genera un reporte a partir de un archivo auxiliar y devuelve los arrays necesarios para el reporte final.
    """
    print('creando columnas de reporte resumen')
        
    aux_data_df = pd.read_excel(aux_path, sheet_name='Index Tablas')
    
    report_resumen_array = aux_data_df['Report_Index'].dropna().to_list()
    report_datails_array = aux_data_df['Report_details_index'].dropna().to_list()
    rotation_report_array = aux_data_df['Rotation_report_index'].dropna().to_list()
    required_columns_data = aux_data_df['Required_final_columns'].dropna().to_list()
    
    print('columnas de reporte resumen creadas')
    
    return report_resumen_array, report_datails_array, rotation_report_array, required_columns_data
    
def insert_data(final_path, required_columns_data, final_report_file):
    
    """
    Inserta los datos necesarios en el archivo final, creando hojas si es necesario.
    """
    print('Moviendo datos a reporte final')

    # Cargar los datos de las hojas necesarias
    data_df = pd.read_excel(final_path, sheet_name='Archivo Final Play Logger', converters={'Id Rev % Ads': str})
    bdd_data_df = pd.read_excel(final_path, sheet_name='BDD Final Revisada')

    # Filtrar columnas necesarias
    data_df = data_df[required_columns_data]

    sheet_name = 'Detalle Revision'
    bdd_sheet_name = 'BDD Revision'

    if not os.path.exists(final_report_file):
        # Si el archivo no existe, crearlo con ambas hojas
        with pd.ExcelWriter(final_report_file, engine='openpyxl') as writer:
            data_df.to_excel(writer, sheet_name=sheet_name, index=False)
            bdd_data_df.to_excel(writer, sheet_name=bdd_sheet_name, index=False)
    else:
        # Si el archivo ya existe, asegurarnos de que las hojas existan antes de escribir
        with pd.ExcelWriter(final_report_file, engine='openpyxl', mode='a') as writer:
            # Cargar el archivo para verificar sus hojas
            workbook = load_workbook(final_report_file)

            # Si las hojas existen, eliminarlas antes de escribir
            if sheet_name in workbook.sheetnames:
                del workbook[sheet_name]
            if bdd_sheet_name in workbook.sheetnames:
                del workbook[bdd_sheet_name]
            workbook.save(final_report_file)

            # Ahora escribir los nuevos datos
            data_df.to_excel(writer, sheet_name=sheet_name, index=False)
            bdd_data_df.to_excel(writer, sheet_name=bdd_sheet_name, index=False)
        
        print(f"Datos actualizados en {final_report_file}")

def get_vendor_mapping():
    """
    Devuelve un diccionario de mapeo de proveedores para renombrar en el reporte final.
    """
    vendor_mapping = {
            'CC MEDIOS USA LLC': 'CC Medios',
            'NBCUniversal Networks International Spanish Latin America LLC': 'NBC Univ',
            'Sony Pictures Television Advertising Sales Company': 'Sony Pictures',
            'VC MEdios Latin America, LLC   (IntL)': 'VC Medios',
            'INVERCORP LIMITED': 'Invercorp',
            'Buena Vista International, INC': 'Buena Vista International'
        }
    return vendor_mapping

def apply_styles_to_sheet(workbook, sheet_name, table_positions):
    """
    Aplica estilos a las tablas dentro de una hoja específica, según las posiciones dadas.
    """
    ws = workbook[sheet_name]
    
    # Estilos para encabezados
    resumen_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")  # Fondo negro
    resumen_font = Font(color="FFFFFF", bold=True)  # Texto blanco
    detalles_fill = PatternFill(start_color="D35400", end_color="D35400", fill_type="solid")  # Fondo naranja oscuro
    detalles_font = Font(color="FFFFFF", bold=True)  # Texto blanco

    # Estilos para registros
    registros_texto = Font(color="000000")  # Texto negro
    registros_alineacion = Alignment(horizontal="center", vertical="center")  # Centrado
    registros_fondo_blanco = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # Fondo blanco

    # Formatos numéricos
    numero_format = "#,##0.00"  # Formato con miles y 2 decimales
    usd_format = '"$"#,##0.00'  # Formato de moneda USD

    # Aplicar estilos a cada tabla en la hoja
    for start_row, end_row in table_positions:
        # Aplicar estilos al encabezado (Fila start_row, desde columna B hasta antes de U)
        for col_num, cell in enumerate(ws[start_row], 1):  
            col_letter = get_column_letter(col_num)
            if "B" <= col_letter < "T":  
                cell.fill = resumen_fill
                cell.font = resumen_font

        # Aplicar estilos al encabezado de Detalles (Fila start_row, desde columna U en adelante)
        for col_num, cell in enumerate(ws[start_row], 1):  
            col_letter = get_column_letter(col_num)
            if col_letter >= "U":  
                cell.fill = detalles_fill
                cell.font = detalles_font

        # Aplicar estilos a los registros de la tabla de Resumen
        for row in ws.iter_rows(min_row=start_row + 1, max_row=end_row, min_col=2, max_col=19):
            for cell in row:
                cell.font = registros_texto
                cell.fill = registros_fondo_blanco  # Fondo blanco

        # Aplicar estilos a los registros de la tabla de Detalles
        for row in ws.iter_rows(min_row=start_row + 1, max_row=end_row, min_col=21, max_col=ws.max_column):
            for cell in row:
                cell.font = registros_texto
                cell.alignment = registros_alineacion
                cell.fill = registros_fondo_blanco  # Fondo blanco

        # Aplicar formato numérico y moneda solo en las columnas correspondientes
        for row in ws.iter_rows(min_row=start_row + 1, max_row=end_row):
            for cell in row:
                if cell.value and isinstance(cell.value, (int, float)):
                    col_letter = get_column_letter(cell.column)
                    if col_letter == "P":  # P = Formato numérico con miles
                        cell.number_format = numero_format
                    elif col_letter == "S":  # S = Formato de moneda USD
                        cell.number_format = usd_format

def generate_columns(final_report_file, report_resumen_array, report_details_array, rotation_report_array):
    """
    Genera las columnas de resumen y detalles en el archivo final.
    """
    print('Generando columnas de resumen con datos de reporte final')
    
    # Leer datos de la hoja "Detalle Revision"
    df_data = pd.read_excel(final_report_file, sheet_name='Detalle Revision')
    df_bdd = pd.read_excel(final_report_file, sheet_name='BDD Revision')
    
    unique_vendor_array = df_data['Vendor'].dropna().unique()
    vendor_mapping = get_vendor_mapping()
    
    def create_vendor_sheet(unique_vendor_array, final_report_file, vendor_mapping):
        """
        Crea hojas de resumen por proveedor en el archivo final.
        """
        print('Creando hojas de resumen por proveedor')
        wb = load_workbook(final_report_file)
        
        for vendor in unique_vendor_array:
            vendor_name = vendor_mapping.get(vendor, vendor)
            if vendor_name in wb.sheetnames:
                del wb[vendor_name]  # Elimina la hoja si ya existe
            wb.create_sheet(vendor_name)  # Crea la nueva hoja
        
        wb.save(final_report_file)
        print('Hojas de resumen por proveedor creadas')

    def generate_vendor_resume(unique_vendor_array, df_data, BDD_file, final_report_file, vendor_mapping):
        """
        Genera el resumen por proveedor y lo guarda en el archivo final.
        """
        print('Generando resumen por proveedor')
        vendor_set = set(df_data['Vendor'].values)
        
        table_positions_dict = {}
        
        with pd.ExcelWriter(final_report_file, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            for vendor in unique_vendor_array:
                if vendor in vendor_set:
                    vendor_name = vendor_mapping.get(vendor, vendor)
                    vendor_df = df_data[df_data['Vendor'] == vendor]
                    
                    unique_brand_array = vendor_df['Brand'].dropna().unique()
                    
                    start_row = 1  # Fila inicial
                    table_positions = []
                    
                    for brand in unique_brand_array:
                        
                        brand_df = vendor_df[vendor_df['Brand'] == brand]
                        
                        resumen_list, details_list = create_resumen_list(brand_df, BDD_file)
                        df_resumen = pd.DataFrame(resumen_list)
                        df_details = pd.DataFrame(details_list)
                        
                        end_row = start_row + max(len(df_resumen), len(df_details))  # Calcular fin de la tabla
                        table_positions.append(((start_row+1), (end_row+1)))  # Guardar posiciones de la tabla

                        df_resumen.to_excel(writer, sheet_name=vendor_name, startrow=start_row, startcol=1, index=False)                        
                        
                        #añadir la columna de detalles desde la fila 2 columna U
                        df_details.to_excel(writer, sheet_name=vendor_name, startrow=start_row, startcol=20, index=False)
                                                
                        start_row += max(len(df_resumen), len(df_details)) + 3
                        
                        fila_total = start_row - 2
                        
                        columnas_sumar = [
                            'PAID SPOTS IO', 'BONUS SPOTS IO',
                            'SPOT PAID TRANSMITTED', 'SPOTS BONUS TRANSMITTED',
                            'SPOTS PAID RECOGNIZED', 'SPOTS BONUS RECOGNIZED',
                            'SPOT PAID NOT RECOGNIZED', 'SPOT BONUS NOT RECOGNIZED',
                            'SPEND LOCAL CURRENT', 'TOTAL SPEND DOLARIZED'
                        ]
                        # Cargar el archivo y la hoja para escribir las sumas
                        wb = load_workbook(final_report_file)
                        ws = wb[vendor_name]

                        # Estilos a aplicar
                        italic_font = Font(color="000000", italic=True)  # Letra negra en cursiva
                        white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                        number_format = "#,##0"  # Número con separador de miles
                        usd_format = '"$"#,##0.00'  # Moneda USD con separador de miles

                        # Escribir label "TOTAL GENERAL" si existe la columna BRAND
                        if 'BRAND' in df_resumen.columns:
                            col_letter = get_column_letter(df_resumen.columns.get_loc('BRAND') + 2)
                            cell_label = ws[f"{col_letter}{fila_total}"]
                            cell_label.value = 'TOTAL GENERAL'
                            cell_label.font = italic_font
                            cell_label.fill = white_fill

                        # Escribir cada suma en su columna correspondiente con estilos
                        for col in columnas_sumar:
                            if col in df_resumen.columns:
                                suma = df_resumen[col].sum()
                                col_letter = get_column_letter(df_resumen.columns.get_loc(col) + 2)  # +2 porque empieza en columna B
                                cell = ws[f"{col_letter}{fila_total}"]
                                cell.value = suma
                                cell.font = italic_font
                                cell.fill = white_fill

                                # Aplicar formato de número o moneda según la columna
                                if col == 'TOTAL SPEND DOLARIZED':
                                    cell.number_format = usd_format
                                else:
                                    cell.number_format = number_format
                    
                    table_positions_dict[vendor_name] = table_positions
        
        wb = load_workbook(final_report_file)
        for vendor, positions in table_positions_dict.items():
            apply_styles_to_sheet(wb, vendor, positions)
        wb.save(final_report_file)
        
        print(f"Resumen guardado en {final_report_file}")

    def create_resumen_list(vendor_df, BDD_file):
        """
        Crea una lista de resumen y detalles para un proveedor específico.
        """
        unique_feed_index_array = vendor_df['Feed Index'].dropna().unique()
        resumen_list = []
        details_list = []
        unique_brand_array = vendor_df['Brand'].dropna().unique()

        for brand in unique_brand_array:
            for feed_index in unique_feed_index_array:
                sub_df = vendor_df[vendor_df['Feed Index'] == feed_index]
                unique_brand_array = sub_df['Brand'].dropna().unique()
                unique_duration_array = sub_df['Duracion'].dropna().unique()

                for duration in unique_duration_array:
                    feed_country_value = vendor_df.loc[vendor_df['Feed Index'] == feed_index, 'Feed'].dropna().unique()
                    feed_country = ', '.join(feed_country_value.astype(str))
                    sumary, details = create_resumen_row(vendor_df, BDD_file, feed_index, brand, duration, feed_country)
                    resumen_list.append(sumary)
                    details_list.append(details)
        
        
                   
        return resumen_list, details_list

    def create_resumen_row(vendor_df, BDD_file, feed_index, brand, duration, feed_country):
        """
        Crea una fila de resumen y detalles para un proveedor específico.
        """
        sumary = {
            'BRAND': brand,
            'FEED INDEX': feed_index,
            'CHANNEL NAME': ', '.join(vendor_df[vendor_df['Feed Index'] == feed_index]['Channel'].dropna().unique().astype(str)),
            'FEED COUNTRY': feed_country,
            'VENDOR (NETSUIT)': ', '.join(vendor_df[vendor_df['Feed Index'] == feed_index]['Vendor'].dropna().unique().astype(str)),
            'DURATION': duration,
            'PAID SPOTS IO': len(BDD_file[(BDD_file['Feed Index'] == feed_index) & (BDD_file['Brand'] == brand) & (BDD_file['Duration'] == duration) & (BDD_file['Type Spot'] == 'Paid')]),
            'BONUS SPOTS IO': len(BDD_file[(BDD_file['Feed Index'] == feed_index) & (BDD_file['Brand'] == brand) & (BDD_file['Duration'] == duration) & (BDD_file['Type Spot'] == 'Bonus')]),
            'SPOT PAID TRANSMITTED': len(vendor_df[(vendor_df['Feed Index'] == feed_index) & (vendor_df['Brand'] == brand) & (vendor_df['Duracion'] == duration) & (vendor_df['Type Spot'] == 'Paid')]),
            'SPOTS BONUS TRANSMITTED': len(vendor_df[(vendor_df['Feed Index'] == feed_index) & (vendor_df['Brand'] == brand) & (vendor_df['Duracion'] == duration) & (vendor_df['Type Spot'] == 'Bonus')]),
            'SPOTS PAID RECOGNIZED': len(vendor_df[(vendor_df['Feed Index'] == feed_index) & (vendor_df['Brand'] == brand) & (vendor_df['Duracion'] == duration) & (vendor_df['Type Spot'] == 'Paid') & (vendor_df['Final Result'] == 'Ok')]),
            'SPOTS BONUS RECOGNIZED': len(vendor_df[(vendor_df['Feed Index'] == feed_index) & (vendor_df['Brand'] == brand) & (vendor_df['Duracion'] == duration) & (vendor_df['Type Spot'] == 'Bonus') & (vendor_df['Final Result'] == 'Ok')]),
            'SPOT PAID NOT RECOGNIZED': len(vendor_df[(vendor_df['Feed Index'] == feed_index) & (vendor_df['Brand'] == brand) & (vendor_df['Duracion'] == duration) & (vendor_df['Type Spot'] == 'Paid') & (vendor_df['Final Result'] == 'No')]),
            'SPOT BONUS NOT RECOGNIZED': len(vendor_df[(vendor_df['Feed Index'] == feed_index) & (vendor_df['Brand'] == brand) & (vendor_df['Duracion'] == duration) & (vendor_df['Type Spot'] == 'Bonus') & (vendor_df['Final Result'] == 'No')]),
            'SPEND LOCAL CURRENT': vendor_df[(vendor_df['Feed Index'] == feed_index) & (vendor_df['Brand'] == brand) & (vendor_df['Duracion'] == duration) & (vendor_df['Final Result'] == 'Ok')]['Rate'].sum(),
            'TYPE SPOT': 'Paid' if 'Paid' in vendor_df['Type Spot'].values or 'Bonus' in vendor_df['Type Spot'].values else ('Cost Zero' if 'Cost Zero' in vendor_df['Type Spot'].values else 'Bonus'),
            'OBSERVATIONS': feed_country,
            'TOTAL SPEND DOLARIZED': vendor_df[(vendor_df['Feed Index'] == feed_index) & (vendor_df['Brand'] == brand) & (vendor_df['Final Result'] == 'Ok')]['Rate'].sum(),

            }
        
        details = {
            'SPOT DUPLICADO': len(vendor_df[(vendor_df['Feed Index'] == feed_index) & (vendor_df['Brand'] == brand) & (vendor_df['Duracion'] == duration)  & (vendor_df['Spot Observation'] == 'Spot Duplicado')]),
            'SPOT NO SOLICITADO': len(vendor_df[(vendor_df['Feed Index'] == feed_index) & (vendor_df['Brand'] == brand) & (vendor_df['Duracion'] == duration) & (vendor_df['Spot Observation'] == 'Spot No solicitado')]),
            'CREATIVO INCORRECTO': len(vendor_df[(vendor_df['Feed Index'] == feed_index) & (vendor_df['Brand'] == brand) & (vendor_df['Duracion'] == duration) & (vendor_df['Creative observation'] == 'Creativo incorrecto')]),
            'CREATIVO TRANSMITIDO INCORRECTAMENTE': len(vendor_df[(vendor_df['Feed Index'] == feed_index) & (vendor_df['Brand'] == brand) & (vendor_df['Duracion'] == duration) & (vendor_df['Creative observation'] == 'Creativo transmitido incorrectamente')]),
            'BACK TO BACK': len(vendor_df[(vendor_df['Feed Index'] == feed_index) & (vendor_df['Brand'] == brand) & (vendor_df['Duracion'] == duration) & (vendor_df['Back to back'] == 'Back to back')]),
        }
        
        return sumary, details
    
    create_vendor_sheet(unique_vendor_array, final_report_file, vendor_mapping)
    generate_vendor_resume(unique_vendor_array, df_data, df_bdd, final_report_file, vendor_mapping)

def generate_rotation_tables(final_revision_path, aux_path):
    """
    Genera las tablas de rotación para cada proveedor en el archivo final.
    """
    vendor_mapping = get_vendor_mapping()
    
    month_rotation_df = pd.read_excel(aux_path, sheet_name='Month_Rotation')
    revision_details_df = pd.read_excel(final_revision_path, sheet_name='Detalle Revision')
    
    revision_details_df['Date Time Zone'] = pd.to_datetime(revision_details_df['Date Time Zone'], errors='coerce')
    month_rotation_df['Start date'] = pd.to_datetime(month_rotation_df['Start date'], errors='coerce')
    month_rotation_df['End date'] = pd.to_datetime(month_rotation_df['End date'], errors='coerce')
        
    unique_vendors = revision_details_df['Vendor'].unique()
    
    
    with pd.ExcelWriter(final_revision_path, mode='a', if_sheet_exists='overlay') as writer:
        for vendor in unique_vendors:
            
            vendor_filtered_df = revision_details_df[revision_details_df['Vendor'] == vendor]
            sheet_name = vendor_mapping.get(vendor, vendor)
            
            #abrir la hoja y detectar el tamaño de la tabla resumen para el vendor, y agurdar la longitud en una variable
            workbook = load_workbook(final_revision_path)
            worksheet = workbook[sheet_name]
            max_row = worksheet.max_row
            start_row = max_row + 3
            workbook.close()
            
            unique_feed_index = vendor_filtered_df['Feed Index'].unique()
            
            for feed_index in unique_feed_index:
                channel_name = vendor_filtered_df.loc[vendor_filtered_df['Feed Index'] == feed_index, 'Channel'].values[0]
                feed_index_filtered_df = vendor_filtered_df[vendor_filtered_df['Feed Index'] == feed_index]          
                feed_country_value = next(iter(vendor_filtered_df.loc[vendor_filtered_df['Feed Index'] == feed_index, 'Feed'].dropna().unique()), "")                
                
                table_data = []
                
                for id_rev in month_rotation_df['Id Rev % Ads'].unique():
                    fecha_ads_id = month_rotation_df.loc[month_rotation_df['Id Rev % Ads'] == id_rev, 'Id Fecha Ads'].values[0]
                    start_date = month_rotation_df.loc[month_rotation_df['Id Rev % Ads'] == id_rev, 'Start date'].values[0]
                    end_date = month_rotation_df.loc[month_rotation_df['Id Rev % Ads'] == id_rev, 'End date'].values[0]
                    current_ad_brand = month_rotation_df.loc[month_rotation_df['Id Rev % Ads'] == id_rev, 'Brand'].values[0]
                    
                    date_filtered_df= feed_index_filtered_df[feed_index_filtered_df['Date Time Zone'].between(start_date, end_date) & (feed_index_filtered_df['Brand'] == current_ad_brand)]
                    
                    #Filtrar el dataframe por el id rev % ads
                    relevant_ads = date_filtered_df[(date_filtered_df['Id Rev % Ads'] == id_rev) & (date_filtered_df['Final Result'] == 'Ok')]
                    
                    total_ads = date_filtered_df[date_filtered_df['Final Result'] == 'Ok']
                    
                    expected_percentage = (month_rotation_df.loc[month_rotation_df['Id Rev % Ads'] == id_rev, 'Percentage'].values[0])*100
                    real_percentage = (relevant_ads.shape[0] / total_ads.shape[0])*100 if total_ads.shape[0] > 0 else 0
                    
                    #leave only 1 decimak value on real percentage
                    real_percentage = round(real_percentage, 1)
                    diff_pp = real_percentage - expected_percentage
                    diff_pp = round(diff_pp, 1)
                    
                    row_data = {
                        'Id Fechas Ads': fecha_ads_id,
                        'Id Rev % Ads': id_rev,
                        'Start Date': pd.Timestamp(start_date).strftime('%m/%d/%Y %H:%M:%S') if pd.notna(start_date) else '',
                        'End Date': pd.Timestamp(end_date).strftime('%m/%d/%Y %H:%M:%S') if pd.notna(end_date) else '',
                        'Ad': month_rotation_df.loc[month_rotation_df['Id Rev % Ads'] == id_rev, 'Creativo'].values[0],
                        'Brand': month_rotation_df.loc[month_rotation_df['Id Rev % Ads'] == id_rev, 'Brand'].values[0],
                        '# Ads': total_ads[(total_ads['Id Rev % Ads'] == id_rev)].shape[0],
                        '% Esperado': expected_percentage,
                        '% Real': real_percentage,
                        'Diff p.p': f"{diff_pp}pp"
                    }
                    table_data.append(row_data)
                
                table_df = pd.DataFrame(table_data)
                table_df.to_excel(writer, sheet_name=sheet_name, startrow=start_row, startcol=1, index=False)
                
                worksheet = writer.book[sheet_name]
                # Merge cells for title
                worksheet.merge_cells(start_row=start_row, start_column=2, end_row=start_row, end_column=11)
                # Aplicar formato al título (Channel Name - Feed Index)
                title_cell = worksheet.cell(row=start_row, column=2, value=f"{channel_name} - {feed_country_value} - {feed_index}")
                title_cell.fill = PatternFill(start_color='000000', fill_type='solid')  # Negro
                title_cell.font = Font(color='FFFFFF', bold=True)  # Blanco y negrita
                
                start_row += 1
                
                # Aplicar formato a encabezados
                header_fill = PatternFill(start_color='000080', fill_type='solid')  # Azul oscuro
                header_font = Font(color='FFFFFF', bold=True)
                for col in range(2, 12):
                    cell = worksheet.cell(row=start_row, column=col)
                    cell.fill = header_fill
                    cell.font = header_font
                    
                start_row += 1
                
                # Aplicar formato a registros
                border = Border(bottom=Side(style='dotted', color='000000'))  # Línea divisoria punteada negra
                
                previous_brand = None
                for index, row in table_df.iterrows():
                    row_fill = PatternFill(start_color='D3D3D3' if row['Id Fechas Ads'] % 2 != 0 else 'FFFFFF', fill_type='solid')
                    
                    for col, value in enumerate(row, start=2):
                        cell = worksheet.cell(row=start_row, column=col, value=value)
                        cell.fill = row_fill
                        cell.alignment = Alignment(horizontal='center')  # Centrar desde '# Ads'
                    
                    if previous_brand and previous_brand != row['Brand']:
                        for col in range(2, 11):
                            worksheet.cell(row=start_row - 1, column=col+1).border = border
                    previous_brand = row['Brand']
                    start_row += 1
                start_row += 2
    
def set_column_widths(file_path):
    """
    Ajusta el ancho de las columnas en el archivo Excel.
    A las hojas no excluidas se les pone en las columnas A y T un ancho fijo de 2,
    y se autoajustan las demás columnas.
    A las hojas excluidas se les hace autofit general.
    """
    workbook = load_workbook(file_path)
    excluded_sheets = ['Detalle Revision', 'BDD Revision']

    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]

        if sheet_name not in excluded_sheets:
            worksheet.column_dimensions['A'].width = 2
            worksheet.column_dimensions['T'].width = 2

            for column_cells in worksheet.columns:
                col_idx = column_cells[0].column
                col_letter = get_column_letter(col_idx)

                if col_letter in ['A', 'T']:
                    continue

                max_length = 0
                for cell in column_cells:
                    try:
                        if cell.value is not None:
                            max_length = max(max_length, len(str(cell.value)))
                    except Exception:
                        pass

                adjusted_width = (max_length + 2) * 1.2 if max_length > 0 else 8
                worksheet.column_dimensions[col_letter].width = adjusted_width

        else:
            for column_cells in worksheet.columns:
                col_idx = column_cells[0].column
                col_letter = get_column_letter(col_idx)

                max_length = 0
                for cell in column_cells:
                    try:
                        if cell.value is not None:
                            max_length = max(max_length, len(str(cell.value)))
                    except Exception:
                        pass

                adjusted_width = (max_length + 2) * 1.2 if max_length > 0 else 8
                worksheet.column_dimensions[col_letter].width = adjusted_width

    # Guarda el archivo con los cambios
    workbook.save(file_path)
    workbook.close()
    
def generate_schedule_sheets(final_report_file, aux_path):
    """
    Genera hojas de programación para cada proveedor en el archivo final.
    """
    # Cargar los datos desde el archivo Excel
    df_data = pd.read_excel(final_report_file, sheet_name='Detalle Revision')
    df_bdd_data = pd.read_excel(final_report_file, sheet_name='BDD Revision')
    df_datos_franjas = pd.read_excel(aux_path, sheet_name='Channel Info Monitoria')
    
    # Convertir la columna 'Date Time Zone' a tipo datetime
    df_data['Date Time Zone'] = pd.to_datetime(df_data['Date Time Zone'], errors='coerce')
    df_bdd_data['Date Time Zone'] = pd.to_datetime(df_bdd_data['Date Time Zone'], errors='coerce')
    
    vendor_mapping = get_vendor_mapping()
    unique_vendor_array = df_data['Vendor'].dropna().unique()
    
    min_day = df_data['Date Time Zone'].min().day
    max_day = df_data['Date Time Zone'].max().day
    
    
    possible_dayparts = {'Madrugada', 'Morning', 'Afternoon', 'Prime time'}
    # Posiciones iniciales
    start_col = 2  # Columna inicial para el título
    first_data_col = 3  # Primera columna de datos (E en Excel)
    
    with pd.ExcelWriter(final_report_file, engine='openpyxl', mode='a', if_sheet_exists="overlay") as writer:
        for vendor in unique_vendor_array:
            vendor_name = vendor_mapping.get(vendor, vendor)
            
            vendor_df = df_data[df_data['Vendor'] == vendor]
            vendor_bdd_df = df_bdd_data[df_bdd_data['Vendor'] == vendor]
            
            unique_feed_index = vendor_df['Feed Index'].dropna().unique()
            sheet_name = f"Grid - {vendor_name}"[:31]
            
            # Crear hoja si no existe, si existe eliminarla y crear una nueva
            if sheet_name in writer.book.sheetnames:
                del writer.book[sheet_name]
            writer.book.create_sheet(sheet_name)
            worksheet = writer.book[sheet_name]
            
            row_offset = 2  # Control de posición vertical para cada tabla
            
            for feed_index in unique_feed_index:
                daypart_dict = {}
                
                feed_index_df = vendor_df[vendor_df['Feed Index'] == feed_index]
                feed_index_bdd_df = vendor_bdd_df[vendor_bdd_df['Feed Index'] == feed_index]
                
                unique_brands = feed_index_df['Brand'].dropna().unique()
                channel_name = next(iter(feed_index_df['Channel'].dropna().unique()), "")
                feed_country_value = next(iter(feed_index_df['Feed'].dropna().unique()), "")
                revision_type = feed_index_df['Revision type'].iloc[0]
                franja_data = df_datos_franjas[df_datos_franjas['Feed Index'] == feed_index].iloc[0]
                daypart_dict[feed_index] = {
                    "Madrugada": franja_data['Start - Madrugada'].hour,
                    "Morning": franja_data['Start - Morning'].hour,
                    "Afternoon": franja_data['Start - Afternoon'].hour,
                    "Prime time": franja_data['Start - Prime Time'].hour
                }
                for brand in unique_brands:
                    brand_df = feed_index_df[feed_index_df['Brand'] == brand]
                    brand_bdd_df = feed_index_bdd_df[feed_index_bdd_df['Brand'] == brand]
                    
                    # Posiciones de la tabla
                    start_row = row_offset
                    date_row = start_row + 1   # Fila donde se colocarán las fechas
                    header_row = start_row + 2  # Fila para los encabezados IOs, CT, ST
                    data_start_row = start_row + 3  # Fila donde comienzan los datos
                    hour_col = 2  # Columna donde se colocará la hora
                    
                    # Escribir el encabezado de la tabla
                    worksheet.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=20)
                    title_cell = worksheet.cell(row=start_row, column=start_col, value=f"{channel_name} - {feed_country_value} - {feed_index} - {brand}")
                    title_cell.fill = PatternFill(start_color='000000', fill_type='solid')
                    title_cell.font = Font(color='FFFFFF', bold=True)
                    
                    header_fill = PatternFill(start_color='000000', fill_type='solid')
                    header_font = Font(color='FFFFFF', bold=True)
                    header_align = Alignment(horizontal='center')

                    # Agregar encabezado "Hr" con estilo
                    hr_cell = worksheet.cell(row=header_row, column=hour_col, value="Hr")
                    hr_cell.fill = header_fill
                    hr_cell.font = header_font
                    hr_cell.alignment = header_align

                    # Estilo para celdas de hora (0–23)
                    for i, hour in enumerate(range(24), start=data_start_row):
                        hour_cell = worksheet.cell(row=i, column=hour_col, value=hour)
                        hour_cell.fill = header_fill  # Negro
                        hour_cell.font = header_font  # Blanco y bold
                        hour_cell.alignment = header_align  # Centrado (opcional)
                    
                    # Procesar cada día
                    col_offset = 0  # Desplazamiento de columnas para los datos
                    for day in range(min_day, max_day + 1):
                        day_df = brand_df[brand_df['Date Time Zone'].dt.day == day]
                        day_bdd_df = brand_bdd_df[brand_bdd_df['Date Time Zone'].dt.day == day]
                        
                        ios_spots = []
                        ct_spots = []
                        st_spots = []
                        
                        if revision_type == 1:
                            #convertir la fecha a datetime en la columna 'Fecha Final Revision'
                            day_df_copy = day_df.copy()
                            day_df_copy['Fecha Final Revision'] = pd.to_datetime(day_df_copy['Fecha Final Revision'], errors='coerce')
                            for hour in range(24):
                                df_ios_spots = day_bdd_df[day_bdd_df['Date Time Zone'].dt.hour == hour]
                                df_ct_spots = day_df_copy[day_df_copy['Fecha Final Revision'].dt.hour == hour]
                                df_st_spots = df_ct_spots[df_ct_spots['Final Result'] == 'Ok']
                                
                                ios_spots.append(df_ios_spots.shape[0])
                                ct_spots.append(df_ct_spots.shape[0])
                                st_spots.append(df_st_spots.shape[0])
                        elif revision_type == 2:
                            ios_spots = [0] * 24  # Aseguramos 24 posiciones con ceros
                            ct_spots = [0] * 24
                            st_spots = [0] * 24

                            for day_part in possible_dayparts:
                                # Filtrar los datos que pertenecen a este day_part
                                start_daypart = daypart_dict[feed_index].get(day_part, None)
                                
                                day_part_df = day_df[day_df['Fecha Final Revision'] == day_part]
                                day_part_bdd_df = day_bdd_df[day_bdd_df['Franja'] == day_part]  # Filtrar por Franja
                                day_part_st_df = day_part_df[day_part_df['Final Result'] == 'Ok']  # Filtrar por 'Final Result' == 'Ok'

                                # Obtener la hora de inicio de este day_part

                                # Verificación de valores
                                if start_daypart is None or not (0 <= start_daypart <= 23):
                                    continue  # Saltar iteración si no hay una hora válida

                                # Iterar en todas las horas (0-23), asegurando que los valores sean 0 en las demás horas
                                for hour in range(24):
                                    if hour == start_daypart:
                                        # En lugar de sobrescribir, ahora sumamos los valores existentes
                                        ios_spots[hour] += day_part_bdd_df.shape[0]  # Total de registros en BDD con la franja
                                        ct_spots[hour] += day_part_df.shape[0]  # Total de registros en CT con la franja
                                        st_spots[hour] += day_part_st_df.shape[0]  # Total de registros en ST con la franja y resultado 'Ok'
                                    else:
                                        # Asegurar que los valores no se desplacen accidentalmente
                                        if hour >= len(ios_spots):
                                            ios_spots.append(0)
                                            ct_spots.append(0)
                                            st_spots.append(0)
                        elif revision_type == 3:
                            ios_spots = [0] * 24  # Aseguramos 24 posiciones con ceros
                            ct_spots = [0] * 24
                            st_spots = [0] * 24
                            
                            total_spots = day_df.shape[0]
                            total_ios_spots = day_bdd_df.shape[0]
                            total_st_spots = day_df[day_df['Final Result'] == 'Ok'].shape[0]

                            # Asignamos los valores de los spots a las 8 AM (índice 8)
                            ct_spots[8] = total_spots
                            ios_spots[8] = total_ios_spots
                            st_spots[8] = total_st_spots
                        
                        # Colocar la fecha en la fila `date_row` alineada con la columna `ST`
                        date_col = first_data_col + col_offset + 2  # ST está 2 columnas después de IOs
                        #Unir celdas de fecha desde 
                        worksheet.merge_cells(start_row=date_row, start_column=date_col - 2, end_row=date_row, end_column=date_col)
                        worksheet.cell(row=date_row, column=date_col-2, value=f"{day}")
                        
                        # Colocar los encabezados IOs, CT, ST en la fila `header_row`
                        worksheet.cell(row=header_row, column=date_col - 2, value="IOs")
                        worksheet.cell(row=header_row, column=date_col - 1, value="CT")
                        worksheet.cell(row=header_row, column=date_col, value="ST")
                        
                        # Estilo para encabezados
                        header_fill = PatternFill(start_color='000000', fill_type='solid')
                        header_font = Font(color='FFFFFF', bold=True)

                        # Aplicar estilo a los encabezados IOs, CT, ST
                        for col in range(date_col - 2, date_col + 1):
                            cell = worksheet.cell(row=header_row, column=col)
                            cell.fill = header_fill
                            cell.font = header_font

                        # Estilo para la celda de la fecha
                        date_cell = worksheet.cell(row=date_row, column=date_col - 2)
                        date_cell.fill = header_fill
                        date_cell.font = header_font
                        
                        # Colocar los datos de IOs, CT y ST
                        for i, (ios, ct, st) in enumerate(zip(ios_spots, ct_spots, st_spots), start=data_start_row):
                            # Estilos
                            gray_fill = PatternFill(start_color='D7D7D9', fill_type='solid')
                            white_fill = PatternFill(start_color='FFFFFF', fill_type='solid')

                            # IOs
                            cell_ios = worksheet.cell(row=i, column=date_col - 2, value=ios)
                            cell_ios.fill = white_fill

                            # CT
                            cell_ct = worksheet.cell(row=i, column=date_col - 1, value=ct)
                            cell_ct.fill = white_fill

                            # ST
                            cell_st = worksheet.cell(row=i, column=date_col, value=st)
                            cell_st.fill = gray_fill
                        
                        col_offset += 3  # Moverse 3 columnas a la derecha por cada día
                    
                    row_offset += 30  # Espacio entre tablas para el mismo proveedor
            
            for col in worksheet.columns:
                col_letter = col[0].column_letter  # Obtener la letra de la columna
                worksheet.column_dimensions[col_letter].width = 3  # Ajustar el ancho de la columna a 30   
            writer.book.save(final_report_file)
                            
def full_report(aux_path, final_path, final_report_file, log_func=None):
    """
    Genera el reporte final a partir de los datos auxiliares y el archivo final.
    """
    if log_func: log_func("🗑️ Eliminando reporte final anterior si existe...")
    if os.path.exists(final_report_file):
        os.remove(final_report_file)

    if log_func: log_func("📊 Generando datos del reporte...")
    report_resumen_array, report_datails_array, rotation_report_array, required_columns_data = generate_report(aux_path)

    if log_func: log_func("📥 Insertando datos al reporte final...")
    insert_data(final_path, required_columns_data, final_report_file)

    if log_func: log_func("🧱 Generando columnas de resumen y detalle...")
    generate_columns(final_report_file, report_resumen_array, report_datails_array, rotation_report_array)

    if log_func: log_func("🔁 Generando tablas de rotación...")
    generate_rotation_tables(final_report_file, aux_path)

    if log_func: log_func("📐 Ajustando ancho de columnas...")
    set_column_widths(final_report_file)

    if log_func: log_func("🗓️ Generando hojas de programación...")
    generate_schedule_sheets(final_report_file, aux_path)

    if log_func: log_func("✅ Reporte final generado correctamente.")
