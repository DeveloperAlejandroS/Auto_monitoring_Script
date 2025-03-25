import pandas as pd
from openpyxl import *
import pytz
import os


def get_revision_conditions(excel_path, aux_path, sheet_name):
    # Leer los datos del archivo principal y auxiliar
    df = pd.read_excel(excel_path, sheet_name=sheet_name)
    df_auxiliar = pd.read_excel(aux_path, sheet_name='Channel Info Monitoria')
    
    # Asegurarse de que Feed Index esté en formato de texto en ambos DataFrames
    df['Feed Index'] = df['Feed Index'].astype(str).str.strip()
    df_auxiliar['Feed Index'] = df_auxiliar['Feed Index'].astype(str).str.strip()
    
    df_auxiliar_grouped = df_auxiliar.groupby('Feed Index', as_index=False).agg({
        'Channel': ', '.join,  # Concatenar los valores de Channel
        'Condition +/-': 'first',  # Tomar el primer valor
        'Revision type': 'first',  # Tomar el primer valor
        'Rotation Id': 'first',
        'Time Zone CT&HT': 'first',
        'Time Zone IO\'s': 'first'
    })
    
    # Crear un diccionario donde la clave es 'Feed Index' y el valor es una lista con los datos correspondientes
    aux_dict = df_auxiliar_grouped.set_index('Feed Index').to_dict(orient='index')
    
    # Crear DataFrame vacío para almacenar los resultados
    result_data = {
        'Condition +/-': [],
        'Revision type': [],
        'Rotation Id': [],
        'Time Zone CT&HT': [],
        'Time Zone IO\'s': []
    }
    
    # Iterar sobre las filas del DataFrame principal
    for i, row in df.iterrows():
        feed_index = row['Feed Index']
        
        # Si el 'Feed Index' está en el diccionario auxiliar, agregar los datos correspondientes
        if feed_index in aux_dict:
            aux_data = aux_dict[feed_index]
            result_data['Condition +/-'].append(aux_data.get('Condition +/-', None))
            result_data['Revision type'].append(aux_data.get('Revision type', None))
            result_data['Rotation Id'].append(aux_data.get('Rotation Id', None))
            result_data['Time Zone CT&HT'].append(aux_data.get('Time Zone CT&HT', None))
            result_data['Time Zone IO\'s'].append(aux_data.get('Time Zone IO\'s', None))
        else:
            # Si no se encuentra el 'Feed Index', agregar valores NaN
            result_data['Condition +/-'].append(None)
            result_data['Revision type'].append(None)
            result_data['Rotation Id'].append(None)
            result_data['Time Zone CT&HT'].append(None)
            result_data['Time Zone IO\'s'].append(None)
    
    # Convertir los resultados en un DataFrame
    df_conditions = pd.DataFrame(result_data)
    
    # Validar que el número de filas en el resultado sea consistente con el archivo original
    if len(df_conditions) != len(df):
        raise ValueError("El número de filas en el resultado no coincide con el archivo original.")
    
    # Escribir el resultado en el archivo Excel en la misma hoja
    with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        df_conditions.to_excel(writer, sheet_name=sheet_name, startrow=0, startcol=12, index=False)
    
def get_date_rev(excel_path, sheet_name):
    df = pd.read_excel(excel_path, sheet_name=sheet_name)
    
    # Asegurarse de que las columnas 'Fecha' y 'Horario' estén en el formato adecuado
    df['Fecha'] = pd.to_datetime(df['Fecha'], errors='coerce').dt.strftime('%m/%d/%Y')  # Formato de fecha m/d/Y
    df['Horario'] = pd.to_datetime(df['Horario'], format='%H:%M:%S', errors='coerce').dt.strftime('%H:%M:%S')  # Formato de hora H:M:S
    
    # Concatenar la fecha y la hora en una nueva columna 'Date Rev'
    df['Date Rev'] = df['Fecha'] + ' ' + df['Horario']
    
    
    with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        df['Date Rev'].to_excel(writer, sheet_name=sheet_name, startrow=0, startcol=17, index=False)
    
def convert_time_zone(excel_path, aux_path, sheet_name):
    df = pd.read_excel(excel_path, sheet_name=sheet_name)
    df_aux = pd.read_excel(aux_path, sheet_name='Zona Horaria')
    
    # Selección y preparación de zonas horarias
    df_aux = df_aux[['Country', 'Time Zone']]
    dict_aux = df_aux.set_index('Country').to_dict()['Time Zone']
    
    # Convertir 'Date Rev' a datetime
    df['Date Rev'] = pd.to_datetime(df['Date Rev'], errors='coerce')
    
    # Verificar valores no convertidos
    if df['Date Rev'].isna().any():
        print("Algunas fechas no se pudieron convertir. Verifica los datos.")
    
    # Inicializar nueva columna
    df['Date Time Zone'] = None
    
    for idx, row in df.iterrows():
        cert_time = row['Time Zone CT&HT']
        io_time = row['Time Zone IO\'s']
        
        if cert_time in dict_aux and io_time in dict_aux:
            try:
                # Localizar y convertir zona horaria
                localized_time = row['Date Rev'].tz_localize(dict_aux[cert_time])
                converted_time = localized_time.tz_convert(dict_aux[io_time])
                
                # Guardar resultado en la nueva columna
                df.at[idx, 'Date Time Zone'] = converted_time.strftime('%m/%d/%Y %H:%M:%S')
            except Exception as e:
                print(f"Error procesando la fila {idx}: {e}")
    
    # Guardar resultados en el archivo Excel
    with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        df['Date Time Zone'].to_excel(writer, sheet_name=sheet_name, startrow=0, startcol=18, index=False)
            
def gen_DTZ_condition(excel_path, aux_path ,sheet_name):
    
    df = pd.read_excel(excel_path, sheet_name=sheet_name)
    #Convert 'Date Rev' to datetime with format 'm/d/Y H:M:S'
    df['Date Rev'] = pd.to_datetime(df['Date Rev'], errors='coerce')
    df_aux = pd.read_excel(aux_path, sheet_name='Channel Info Monitoria')
    
    df['Date Time Zone - Minutes'] = None
    df['Date Time Zone = Minutes'] = None
    df['Date Time Zone + Minutes'] = None
    df['Day Part - Minutes'] = None
    df['Day Part = Minutes'] = None
    df['Day Part + Minutes'] = None
    df['Full Day - Minutes'] = None
    df['Full Day = Minutes'] = None
    df['Full Day + Minutes'] = None
    
    df_aux = df_aux[['Feed Index','Start - Madrugada','End - Madrugada','Start - Morning','End - Morning','Start - Afternoon','End - Afternoon','Start - Prime Time','End - Prime Time']]
    
    df_auxiliar_grouped = df_aux.groupby('Feed Index', as_index=False).agg({
        'Start - Madrugada': 'first',
        'End - Madrugada': 'first',
        'Start - Morning': 'first',
        'End - Morning': 'first',
        'Start - Afternoon': 'first',
        'End - Afternoon': 'first',
        'Start - Prime Time': 'first',
        'End - Prime Time': 'first'
    })
    
    dic_daypart = df_auxiliar_grouped.set_index('Feed Index').to_dict(orient='index')
    
    # Crear un diccionario donde la clave es 'Feed Index' y el valor es una lista con los datos correspondientes
    aux_dict = df_auxiliar_grouped.set_index('Feed Index').to_dict(orient='index')
    
    df_aux = pd.read_excel(aux_path, sheet_name='Channel Info Monitoria')
    
    for inx, row in df.iterrows():
        
        current_date = row['Date Time Zone']
        current_date = pd.to_datetime(current_date, errors='coerce')
        current_condition = row['Condition +/-']
        
        if current_date is not pd.NaT:
            minus_date = current_date - pd.Timedelta(current_condition, unit='m')
            minus_hour = minus_date.time()
            minus_full_day = minus_date.replace(hour=0, minute=0, second=0)
            
            df.at[inx, 'Date Time Zone - Minutes'] = minus_date.replace(minute=0, second=0).strftime('%m/%d/%Y %H:%M:%S')
            df.at[inx, 'Full Day - Minutes'] = minus_full_day.strftime('%m/%d/%Y %H:%M')
            
            if minus_hour >= dic_daypart[row['Feed Index']]['Start - Madrugada'] and minus_hour < dic_daypart[row['Feed Index']]['End - Madrugada']:
                df.at[inx, 'Day Part - Minutes'] = 'Madrugada'
            elif minus_hour >= dic_daypart[row['Feed Index']]['Start - Morning'] and minus_hour < dic_daypart[row['Feed Index']]['End - Morning']:
                df.at[inx, 'Day Part - Minutes'] = 'Morning'
            elif minus_hour >= dic_daypart[row['Feed Index']]['Start - Afternoon'] and minus_hour < dic_daypart[row['Feed Index']]['End - Afternoon']:
                df.at[inx, 'Day Part - Minutes'] = 'Afternoon'
            elif minus_hour >= dic_daypart[row['Feed Index']]['Start - Prime Time'] and minus_hour < dic_daypart[row['Feed Index']]['End - Prime Time']:
                df.at[inx, 'Day Part - Minutes'] = 'Prime Time'
            
            
            equal_date = current_date.replace(minute=0, second=0)
            equal_hour = equal_date.time()
            equal_full_day = equal_date.replace(hour=0, minute=0, second=0)
            
            df.at[inx, 'Date Time Zone = Minutes'] = equal_date.strftime('%m/%d/%Y %H:%M:%S')
            df.at[inx, 'Full Day = Minutes'] = equal_full_day.strftime('%m/%d/%Y %H:%M')
            
            if equal_hour >= dic_daypart[row['Feed Index']]['Start - Madrugada'] and equal_hour < dic_daypart[row['Feed Index']]['End - Madrugada']:
                df.at[inx, 'Day Part = Minutes'] = 'Madrugada'
            elif equal_hour >= dic_daypart[row['Feed Index']]['Start - Morning'] and equal_hour < dic_daypart[row['Feed Index']]['End - Morning']:
                df.at[inx, 'Day Part = Minutes'] = 'Morning'
            elif equal_hour >= dic_daypart[row['Feed Index']]['Start - Afternoon'] and equal_hour < dic_daypart[row['Feed Index']]['End - Afternoon']:
                df.at[inx, 'Day Part = Minutes'] = 'Afternoon'
            elif equal_hour >= dic_daypart[row['Feed Index']]['Start - Prime Time'] and equal_hour < dic_daypart[row['Feed Index']]['End - Prime Time']:
                df.at[inx, 'Day Part = Minutes'] = 'Prime Time'
            
            plus_date = current_date + pd.Timedelta(current_condition, unit='m')
            plus_hour = plus_date.time()
            plus_full_day = plus_date.replace(hour=0, minute=0, second=0)
            
            df.at[inx, 'Date Time Zone + Minutes'] = plus_date.replace(minute=0, second=0).strftime('%m/%d/%Y %H:%M:%S')
            df.at[inx, 'Full Day + Minutes'] = plus_full_day.strftime('%m/%d/%Y %H:%M')
            
            if plus_hour >= dic_daypart[row['Feed Index']]['Start - Madrugada'] and plus_hour < dic_daypart[row['Feed Index']]['End - Madrugada']:
                df.at[inx, 'Day Part + Minutes'] = 'Madrugada'
            elif plus_hour >= dic_daypart[row['Feed Index']]['Start - Morning'] and plus_hour < dic_daypart[row['Feed Index']]['End - Morning']:
                df.at[inx, 'Day Part + Minutes'] = 'Morning'
            elif plus_hour >= dic_daypart[row['Feed Index']]['Start - Afternoon'] and plus_hour < dic_daypart[row['Feed Index']]['End - Afternoon']:
                df.at[inx, 'Day Part + Minutes'] = 'Afternoon'
            elif plus_hour >= dic_daypart[row['Feed Index']]['Start - Prime Time'] and plus_hour < dic_daypart[row['Feed Index']]['End - Prime Time']:
                df.at[inx, 'Day Part + Minutes'] = 'Prime Time'
            
            
    with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        df['Date Time Zone - Minutes'].to_excel(writer, sheet_name=sheet_name, startrow=0, startcol=19, index=False)
        df['Date Time Zone = Minutes'].to_excel(writer, sheet_name=sheet_name, startrow=0, startcol=20, index=False)
        df['Date Time Zone + Minutes'].to_excel(writer, sheet_name=sheet_name, startrow=0, startcol=21, index=False)
        df['Day Part - Minutes'].to_excel(writer, sheet_name=sheet_name, startrow=0, startcol=22, index=False)
        df['Day Part = Minutes'].to_excel(writer, sheet_name=sheet_name, startrow=0, startcol=23, index=False)
        df['Day Part + Minutes'].to_excel(writer, sheet_name=sheet_name, startrow=0, startcol=24, index=False)
        df['Full Day - Minutes'].to_excel(writer, sheet_name=sheet_name, startrow=0, startcol=25, index=False)
        df['Full Day = Minutes'].to_excel(writer, sheet_name=sheet_name, startrow=0, startcol=26, index=False)
        df['Full Day + Minutes'].to_excel(writer, sheet_name=sheet_name, startrow=0, startcol=27, index=False)


def copydf_to_final_path(excel_path, final_path, sheet_name):
    # Leer el DataFrame desde la hoja especificada
    df = pd.read_excel(excel_path, sheet_name=sheet_name)
    
    # Verificar si el archivo final ya existe
    if os.path.exists(final_path):
        # Si el archivo existe, abrir en modo de añadir ('a')
        with pd.ExcelWriter(final_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
    else:
        # Si el archivo no existe, crearlo y guardar el DataFrame
        with pd.ExcelWriter(final_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
    
    # Renombrar hojas en el archivo original (excel_path)
    wb = load_workbook(excel_path)
    
    if 'Convertido y procesado' in wb.sheetnames:
        ws = wb['Convertido y procesado']
        wb.remove(ws)
    
    if 'Worksheet' in wb.sheetnames:  # Verificar si la hoja 'Worksheet' existe
        ws = wb['Worksheet']
        ws.title = 'Data Play Logger'
    
    if 'Archivo Final Play Logger' in wb.sheetnames:  # Verificar si la hoja 'Archivo Final Play Logger' existe
        ws = wb['Archivo Final Play Logger']
        ws.title = 'Convertido y procesado'
    
    # Guardar los cambios en el archivo original
    wb.save(excel_path)
    

def fetch_additional_columns(excel_path, aux_path, final_path, sheet_name):
    
    print(final_path)
    get_revision_conditions(excel_path, aux_path, sheet_name)
    get_date_rev(excel_path, sheet_name)
    convert_time_zone(excel_path, aux_path, sheet_name)
    gen_DTZ_condition(excel_path, aux_path ,sheet_name)
    copydf_to_final_path(excel_path, final_path, sheet_name)