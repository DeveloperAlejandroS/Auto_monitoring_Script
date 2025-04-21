import pandas as pd
from openpyxl import load_workbook

def apply_transformations_to_excel_file(excel_path, log_func=None):
    """
    Aplica una serie de transformaciones a un archivo de Excel para limpiar y estructurar sus datos.
    
    Transformaciones realizadas:
    1. Elimina caracteres Unicode no deseados (espacios duros `U+00A0`).
    2. Rellena celdas vacías en las primeras 5 columnas con el valor superior inmediato.
    3. Reemplaza fórmulas en celdas por el último valor conocido.
    4. Elimina filas que contienen la palabra "Conteo" en la columna C ('Días laborados').
    5. Convierte las fechas de la columna B a formato `MM/DD/YYYY` y las almacena en la columna F de una nueva hoja.
    
    Finalmente, guarda los cambios en el mismo archivo Excel.
    
    :param excel_path: Ruta del archivo Excel a procesar.
    """
    wb = load_workbook(excel_path)
    ws = wb.active

    try:
        if log_func: log_func("Borrando caracteres Unicode no deseados")
        replace_unicode_character(ws)

        if log_func: log_func("Llenando celdas vacías con el valor superior")
        fill_empty_cells(ws)

        if log_func: log_func("Reemplazando fórmulas por valores fijos")
        replace_formulas_with_values(ws)

        if log_func: log_func("Eliminando filas innecesarias")
        delete_count_rows(ws)

        if log_func: log_func("Formateando fechas en nueva hoja")
        format_date_column(ws)

        wb.save(excel_path)
        if log_func:
            log_func("✅ Archivo guardado")
        else:
            print(f"Archivo guardado en {excel_path}")

    except Exception as e:
        error_msg = f"❌ Error al procesar el archivo: {e}"
        if log_func:
            log_func(error_msg)
        else:
            print(error_msg)
    finally:
        wb.close()


#===============TRANSFORMATION LOGIC===============#

def replace_unicode_character(ws):
    """
    Reemplaza caracteres Unicode no deseados (`U+00A0`, espacio duro) con espacios vacíos
    en todas las celdas de la hoja de trabajo.
    
    :param ws: Hoja de trabajo de un archivo Excel.
    """
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == '\u00A0':
                cell.value = None

def fill_empty_cells(ws):
    """
    Rellena celdas vacías en las primeras 5 columnas con el valor de la celda superior inmediata.
    
    :param ws: Hoja de trabajo de un archivo Excel.
    """
    for col in ws.iter_cols(min_col=1, max_col=5):
        for row in range(2, ws.max_row + 1):
            if col[row-1].value is None:
                cell_above = col[row-2]
                if cell_above.value is not None:
                    col[row-1].value = cell_above.value

def replace_formulas_with_values(ws):
    """
    Reemplaza las fórmulas en las primeras 5 columnas con su último valor conocido.
    
    :param ws: Hoja de trabajo de un archivo Excel.
    """
    for col in ws.iter_cols(min_col=1, max_col=5):
        last_known_value = None
        for row in range(1, ws.max_row + 1):
            cell = col[row-1]
            if cell.value is not None and cell.data_type != 'f':  # Check if it's not a formula
                last_known_value = cell.value
            elif cell.data_type == 'f':  # Replace formula with last known value
                cell.value = last_known_value

def delete_count_rows(ws):
    """
    Elimina filas en las que la tercera columna ('Días laborados') contenga la palabra "Conteo".
    
    :param ws: Hoja de trabajo de un archivo Excel.
    """
    column_index = 3  # Column 'Días laborados'
    rows_to_delete = []
    for row in range(1, ws.max_row + 1):
        cell_value = ws.cell(row=row, column=column_index).value
        if cell_value and "Conteo" in str(cell_value):
            rows_to_delete.append(row)
    
    for row in reversed(rows_to_delete):
        ws.delete_rows(row)

def format_date_column(excel_path):
    """
    Convierte las fechas de la columna B al formato `MM/DD/YYYY` y las guarda en la columna F
    de una nueva hoja llamada 'Archivo Final Play Logger'.
    
    Si la hoja no existe, se crea automáticamente.
    
    :param ws: Hoja de trabajo de un archivo Excel.
    """
    sheet_name = 'Archivo Final Play Logger'

    try:
        df = pd.read_excel(excel_path, usecols='B', names=['Fecha'])
    except Exception as e:
        print(f"Error al leer el archivo Excel: {e}")
        return
    
    # Convertir fechas al formato MM/DD/YYYY
    df['Fecha'] = pd.to_datetime(df['Fecha'], format='%d/%m/%Y', errors='coerce').dt.strftime('%m/%d/%Y')

    # Cargar el archivo Excel y verificar la existencia de la hoja
    wb = load_workbook(excel_path)
    if sheet_name not in wb.sheetnames:
        wb.create_sheet(title=sheet_name)
        print(f"Hoja '{sheet_name}' creada.")
    ws = wb[sheet_name]

    # Escribir las fechas formateadas en la columna 'F'
    for idx, date_value in enumerate(df['Fecha'].fillna('Fecha Inválida'), start=1):
        ws.cell(row=idx, column=6, value=date_value)  # Columna 6 corresponde a 'F'

    # Guardar el archivo
    try:
        wb.save(excel_path)
        print(f"Archivo guardado correctamente en {excel_path}")
    except Exception as e:
        print(f"Error al guardar el archivo Excel: {e}")
    finally:
        wb.close()